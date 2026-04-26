from __future__ import annotations

import io
from decimal import Decimal
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.schemas.calculation import CalculationLineResponse


def build_template_xlsx(
    sheet_name: str,
    headers: list[str],
    example_row: list[str],
    column_widths: dict[str, int] | None = None,
) -> bytes:
    """Build an .xlsx file with a styled header row and one example row.

    The result is what the user will fill in and re-export as CSV (UTF-8, ;).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel limit

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = (column_widths or {}).get(header, max(14, len(header) + 4))

    for col_idx, value in enumerate(example_row, start=1):
        ws.cell(row=2, column=col_idx, value=value)

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


_HEADER_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_GROUP_FILL = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
_GROUP_FONT = Font(bold=True)
_CENTER = Alignment(horizontal="center", vertical="center")


def _to_float(value: Decimal | None) -> float | None:
    return float(value) if value is not None else None


def build_consumption_xlsx(lines: Iterable[CalculationLineResponse]) -> bytes:
    """Excel de consumo de matéria-prima agrupado por grupo.

    Colunas: GRUPO | CÓDIGO | DESCRIÇÃO | UN1 | QTD UN1 | UN2 | QTD UN2.
    Inclui linhas individuais e subtotal por grupo (somente quando todas
    as MPs do grupo compartilham a mesma UN1; caso contrário deixa em branco
    para evitar somar unidades incompatíveis).
    """
    rms = [
        line for line in lines if line.type == "RAW_MATERIAL"
    ]
    rms.sort(
        key=lambda line: (
            (line.group_name or "￿").lower(),
            line.code,
        )
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Consumo MP"

    headers = ["GRUPO", "CÓDIGO", "DESCRIÇÃO", "UN1", "QTD UN1", "UN2", "QTD UN2"]
    widths = {1: 24, 2: 14, 3: 40, 4: 8, 5: 14, 6: 8, 7: 14}
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = widths[col_idx]

    row_idx = 2
    current_group: tuple[str, list[CalculationLineResponse]] | None = None

    def flush_group():
        nonlocal row_idx
        if not current_group:
            return
        name, items = current_group
        uoms1 = {it.uom for it in items}
        uoms2 = {it.uom2 for it in items if it.uom2}
        total1 = sum((Decimal(it.accumulated_quantity) for it in items), start=Decimal("0"))
        total2 = sum(
            (Decimal(it.quantity2) for it in items if it.quantity2 is not None),
            start=Decimal("0"),
        )
        ws.cell(row=row_idx, column=1, value=f"Total {name}").font = _GROUP_FONT
        if len(uoms1) == 1:
            ws.cell(row=row_idx, column=4, value=next(iter(uoms1))).font = _GROUP_FONT
            cell5 = ws.cell(row=row_idx, column=5, value=_to_float(total1))
            cell5.number_format = "#,##0.000"
            cell5.font = _GROUP_FONT
        if len(uoms2) == 1 and total2 > 0:
            ws.cell(row=row_idx, column=6, value=next(iter(uoms2))).font = _GROUP_FONT
            cell7 = ws.cell(row=row_idx, column=7, value=_to_float(total2))
            cell7.number_format = "#,##0.000"
            cell7.font = _GROUP_FONT
        for col in range(1, 8):
            ws.cell(row=row_idx, column=col).fill = _GROUP_FILL
        row_idx += 1

    for line in rms:
        group_name = line.group_name or "Sem grupo"
        if current_group is None or current_group[0] != group_name:
            flush_group()
            current_group = (group_name, [])
        current_group[1].append(line)

        ws.cell(row=row_idx, column=1, value=group_name)
        ws.cell(row=row_idx, column=2, value=line.code)
        ws.cell(row=row_idx, column=3, value=line.description)
        ws.cell(row=row_idx, column=4, value=line.uom)
        cell5 = ws.cell(row=row_idx, column=5, value=_to_float(line.accumulated_quantity))
        cell5.number_format = "#,##0.000"
        if line.uom2:
            ws.cell(row=row_idx, column=6, value=line.uom2)
        if line.quantity2 is not None:
            cell7 = ws.cell(row=row_idx, column=7, value=_to_float(line.quantity2))
            cell7.number_format = "#,##0.000"
        row_idx += 1

    flush_group()

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
