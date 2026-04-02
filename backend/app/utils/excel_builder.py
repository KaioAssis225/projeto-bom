from __future__ import annotations

from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.core.time import now_sp
from app.domain.bom_calculator import CalculationLine


HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NUMERIC_FORMAT = "#.##0,00"


def decimal_to_excel_float(value: Decimal | None) -> float | None:
    if value is None:
        return None
    return float(value)


class BomExcelBuilder:
    def __init__(
        self,
        *,
        lines: list[CalculationLine],
        params: dict,
        product_data: dict,
        output_dir: Path,
    ) -> None:
        self.lines = lines
        self.params = params
        self.product_data = product_data
        self.output_dir = output_dir

    def build(self) -> str:
        self.output_dir.mkdir(parents=True, exist_ok=True)
        filename = f"BOM_CALC_{now_sp().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = self.output_dir / filename

        workbook = Workbook()
        consolidated = workbook.active
        consolidated.title = "Resultado Consolidado"
        self._build_consolidated_sheet(consolidated)

        parameters = workbook.create_sheet(title="Parâmetros")
        self._build_parameters_sheet(parameters)

        workbook.save(file_path)
        return str(file_path)

    def _build_consolidated_sheet(self, sheet) -> None:
        headers = [
            "Código",
            "Descrição",
            "Tipo Item",
            "Grupo",
            "Unidade",
            "Qtd Bruta",
            "Qtd Convertida",
            "Preço Vigente",
            "Custo Total",
        ]
        sheet.append(headers)

        for cell in sheet[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

        total_qtd_bruta = Decimal("0")
        total_cost = Decimal("0")

        for line in self.lines:
            total_qtd_bruta += line.accumulated_quantity
            total_cost += line.line_cost
            sheet.append(
                [
                    line.code,
                    line.description,
                    line.type,
                    line.group_name or "",
                    line.uom,
                    decimal_to_excel_float(line.accumulated_quantity),
                    decimal_to_excel_float(line.accumulated_quantity),
                    decimal_to_excel_float(line.price),
                    decimal_to_excel_float(line.line_cost),
                ]
            )

        total_row_idx = sheet.max_row + 1
        sheet.append(
            [
                "TOTAL",
                "",
                "",
                "",
                "",
                decimal_to_excel_float(total_qtd_bruta),
                "",
                "",
                decimal_to_excel_float(total_cost),
            ]
        )

        for cell in sheet[total_row_idx]:
            cell.font = Font(bold=True)

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=6, max_col=9):
            for cell in row:
                if cell.value is not None:
                    cell.number_format = NUMERIC_FORMAT

        self._autofit_columns(sheet)

    def _build_parameters_sheet(self, sheet) -> None:
        rows = [
            ("Produto raiz", self.product_data["root_label"]),
            ("Data de referência", self.params["reference_date"]),
            ("Filtro por grupo", self.params.get("material_group") or "Sem filtro"),
            ("Usuário solicitante", self.params["requested_by"]),
            ("Data/hora de geração", self.params["generated_at"]),
            ("Referência de simulação", self.params.get("simulation_reference") or ""),
        ]

        for key, value in rows:
            sheet.append([key, value])

        for cell in sheet["A"]:
            cell.font = Font(bold=True)

        self._autofit_columns(sheet)

    @staticmethod
    def _autofit_columns(sheet) -> None:
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            sheet.column_dimensions[column_letter].width = max_length + 2
